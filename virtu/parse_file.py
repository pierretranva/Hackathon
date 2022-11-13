import pandas as pd
from sklearn.neighbors import KNeighborsClassifier
import csv
import openpyxl


def print_values() -> dict:
    # load excel with its path
    ghgdf = pd.read_csv("ghgpdata2021.xlsx")

    # wrkbk = openpyxl.load_workbook("ghgpdata2021.xlsx")

    # sh = wrkbk.active

    out = {}
    positions = ghgdf[["Latitute", "Longitude"]]

    print(positions)
    # count = 0
    # iterate through excel and display data
    # for i in range(5, sh.max_row + 1):
    #     print("\n")
    #     # print("Row ", i, " data :")
    #     l = []
    #     isZero = False
    #     for j in range(9, 12):
    #         cell_obj = sh.cell(row=i, column=j)
    #         if (cell_obj.value == None or cell_obj.value == 0) and j == 11:
    #             isZero = True
    #         l.append(cell_obj.value)

    #     if not isZero:
    #         out[count] = l
    #         count = count + 1
    return out


print_values()
